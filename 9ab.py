from openpyxl import workbook,load_workbook
import time

book=workbook()
sheet=book.active
sheet['A1']=56
sheet['A2']=43
now=time.strftime("%x")
sheet['A3']=now
book.save("sample2.xlsx")
sheetwb=load_workbook("sample2.xlsx")
sheet=workbook.worksheets [0]
sheet2=workbook.createsheet('demo')
name=["abc","zxc","qwe","asd","dfg"]
salary=["12345","23456","34567","45678","56789"]
sheet2.cell(row=1,column=1).value="name"
sheet2.cell(row=1,column=2).value="salary"
j=2
for i in range(0,5):
    sheet2.cell(row=j,column=1).value=name[i]
    sheet2.cell(row=j,column=2).value=salary[i]
j+=1
print(workbook.active)
sheet.cell(row=6,column=9).value="raju"
sheet['F10']=100
print(workbook.active)
print(workbook.shectname)
workbook.save("sample2.xlsx")
print("reading Row1",sheet['A1'].value)
print("reading Row3",sheet['A3'].value)
cells=sheet['A1':'B3']
for c1,c2 in cells:
    print(c1.value,c2.value)
    
